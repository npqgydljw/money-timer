import os
import requests
import tempfile
import pdfplumber
from docx import Document
from openpyxl import load_workbook

class FileProcessor:
    def __init__(self, download_dir=None):
        self.download_dir = download_dir or tempfile.gettempdir()
        os.makedirs(self.download_dir, exist_ok=True)
    
    def download_file(self, url):
        """下载文件到临时目录"""
        try:
            response = requests.get(url, timeout=30)
            response.raise_for_status()
            
            # 提取文件名
            filename = os.path.basename(url.split('?')[0])
            filepath = os.path.join(self.download_dir, filename)
            
            with open(filepath, 'wb') as f:
                f.write(response.content)
            
            return filepath
        except Exception as e:
            print(f"下载文件 {url} 失败: {str(e)}")
            return None
    
    def extract_text(self, file_path):
        """根据文件类型提取文本"""
        ext = os.path.splitext(file_path)[1].lower()
        
        try:
            if ext == '.pdf':
                return self._extract_pdf_text(file_path)
            elif ext in ['.doc', '.docx']:
                return self._extract_docx_text(file_path)
            elif ext in ['.xls', '.xlsx']:
                return self._extract_excel_text(file_path)
            else:
                print(f"不支持的文件类型: {ext}")
                return None
        except Exception as e:
            print(f"处理文件 {file_path} 失败: {str(e)}")
            return None
    
    def _extract_pdf_text(self, pdf_path):
        """提取PDF文本"""
        text = []
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                page_text = page.extract_text()
                if page_text:
                    text.append(page_text)
        return '\n'.join(text)
    
    def _extract_docx_text(self, docx_path):
        """提取Word文档文本"""
        doc = Document(docx_path)
        text = []
        for paragraph in doc.paragraphs:
            if paragraph.text:
                text.append(paragraph.text)
        return '\n'.join(text)
    
    def _extract_excel_text(self, excel_path):
        """提取Excel文件文本"""
        wb = load_workbook(excel_path)
        text = []
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            text.append(f"工作表: {sheet_name}")
            
            for row in sheet.iter_rows(values_only=True):
                row_text = [str(cell) if cell is not None else '' for cell in row]
                if any(row_text):
                    text.append('\t'.join(row_text))
            
            text.append('')
        
        return '\n'.join(text)
    
    def process_file_links(self, file_links):
        """处理文件链接列表，提取所有文本"""
        file_data = []
        
        for link in file_links:
            print(f"处理文件: {link}")
            filepath = self.download_file(link)
            if filepath:
                text = self.extract_text(filepath)
                if text:
                    file_data.append({
                        'source': link,
                        'content': text
                    })
                # 清理临时文件
                try:
                    os.remove(filepath)
                except:
                    pass
        
        return file_data